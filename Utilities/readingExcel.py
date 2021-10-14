import openpyxl

wb=openpyxl.load_workbook("..//excel//testdata.xlsx")
sheet=wb["LoginTest"]
total_rows=sheet.max_row
total_cols=sheet.max_column

print("Total rows are",str(total_rows))
print("Total columns are",str(total_cols))
print(sheet.cell(row=2,column=1).value)

for rows in range(1,total_rows+1):
    for cols in range(1,total_cols+1):
        print(sheet.cell(row=rows,column=cols).value,end="    ")
    print()
