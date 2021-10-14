import openpyxl

def getRowCount(path,sheetName):
    wb=openpyxl.load_workbook(path)
    sheet=wb[sheetName]
    return sheet.max_row

def getColCount(path,sheetName):
    wb = openpyxl.load_workbook(path)
    sheet = wb[sheetName]
    return sheet.max_column

def getCellData(path,sheetName,rowNum,colNum):
    wb = openpyxl.load_workbook(path)
    sheet = wb[sheetName]
    return sheet.cell(row=rowNum,column=colNum).value

def setCellData(path,sheetName,rowNum,colNum,data):
    wb = openpyxl.load_workbook(path)
    sheet = wb[sheetName]
    sheet.cell(row=rowNum, column=colNum).value=data
    wb.save(path)

