import openpyxl

def get_data(sheetname):
    workbook=openpyxl.load_workbook("..//Excel//testdata.xlsx")
    sheet=workbook[sheetname]
    total_rows=sheet.max_row
    total_cols=sheet.max_column
    mainlist=[]
    for rows in range(2,total_rows+1):
        datalist=[]
        for cols in range(1,total_cols+1):
            data=sheet.cell(row=rows,column=cols).value
            datalist.insert(cols,data)
        mainlist.insert(rows,datalist)
    return mainlist
